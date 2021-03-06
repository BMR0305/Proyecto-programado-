import pygame # Libreria principal para la interfaz grafica 
import sys # Librireria del sistema
import os # Libreria del sistema operativo
import random #Libreria de randomizado
#Libreria que permite interactuar con archivos .xlsx (excel)
from openpyxl import load_workbook 
from openpyxl import Workbook
from datetime import datetime #Libreria que accede a la fecha y hora del sistema
os.environ['SDL_VIDEO_CENTERED'] = '0' #se centra la ventana de juego
pygame.init()
#Crear vnetana
size = (700, 700)
screen = pygame.display.set_mode(size)
pygame.display.set_caption("Advice Machine")
#Reloj
clock=pygame.time.Clock()
tiempo = 15
#cargar archivos excel
ventas = load_workbook("Ventas.xlsx")
ventas_s = ventas.active 
mensajes = load_workbook("Mensajes.xlsx")
mensajes_s = mensajes.active
excel_font = pygame.font.SysFont("Times New Roman", 20)
#Fondos
background = pygame.image.load("Imagenes/Menu.jpg").convert()
fondo = pygame.image.load("Imagenes/Fondo.jpg").convert()
fondo_admin = pygame.image.load("Imagenes/fondo_admin.jpg").convert()
#Botonoes
start = pygame.image.load("Imagenes/Start.png").convert()
start.set_colorkey([0,0,0])
start_p = pygame.image.load("Imagenes/Start_p.png").convert()
start_p.set_colorkey([0,0,0])
admin = pygame.image.load("Imagenes/Admin.png").convert()
admin.set_colorkey([0,0,0])
admin_p = pygame.image.load("Imagenes/Admin_p.png").convert()
admin_p.set_colorkey([0,0,0])
español_v = pygame.image.load("Imagenes/Español_v.png").convert()
español_v.set_colorkey([0,0,0])
español_r = pygame.image.load("Imagenes/Español_r.png").convert()
español_r.set_colorkey([0,0,0])
english_v = pygame.image.load("Imagenes/English_v.png").convert()
english_v.set_colorkey([0,0,0])
english_r = pygame.image.load("Imagenes/English_r.png").convert()
english_r.set_colorkey([0,0,0])
ojo = pygame.image.load("Imagenes/Ojo.png").convert()
ojo.set_colorkey([0,0,0])
ojo_c = pygame.image.load("Imagenes/Ojo_c.png").convert()
ojo_c.set_colorkey([0,0,0])
reset = pygame.image.load("Imagenes/Reset.png").convert()
reset.set_colorkey([0,0,0])
reset_p = pygame.image.load("Imagenes/Reset_p.png").convert()
reset_p.set_colorkey([0,0,0])
report = pygame.image.load("Imagenes/Report.png").convert()
report.set_colorkey([0,0,0])
report_p = pygame.image.load("Imagenes/Report_p.png").convert()
report_p.set_colorkey([0,0,0])
shut_down = pygame.image.load("Imagenes/Shut_down.png").convert()
shut_down.set_colorkey([0,0,0])
shut_down_p = pygame.image.load("Imagenes/Shut_down_p.png").convert()
shut_down_p.set_colorkey([0,0,0])
volver = pygame.image.load("Imagenes/Volver.png").convert()
volver.set_colorkey([0,0,0])
volver_p = pygame.image.load("Imagenes/Volver_p.png").convert()
volver_p.set_colorkey([0,0,0])
#Dispensador imagenes
dispensador1 = pygame.image.load("Imagenes/Dispensador/Dispensador1.JPG").convert()
dispensador2 = pygame.image.load("Imagenes/Dispensador/Dispensador2.JPG").convert()
dispensador3 = pygame.image.load("Imagenes/Dispensador/Dispensador3.JPG").convert()
dispensador4 = pygame.image.load("Imagenes/Dispensador/Dispensador4.JPG").convert()
dispensador5 = pygame.image.load("Imagenes/Dispensador/Dispensador5.JPG").convert()
dispensador6 = pygame.image.load("Imagenes/Dispensador/Dispensador6.JPG").convert()
dispensador7 = pygame.image.load("Imagenes/Dispensador/Dispensador7.JPG").convert()
dispensador8 = pygame.image.load("Imagenes/Dispensador/Dispensador8.JPG").convert()
#Sonidos
beep = pygame.mixer.Sound("Sonidos/Boton.wav")
beep.set_volume(0.3)
cerradura = pygame.mixer.Sound("Sonidos/Cerradura.wav")
cerradura.set_volume(0.4)
moneda_s = pygame.mixer.Sound("Sonidos/Moneda.wav")
moneda_s.set_volume(1)
imprimir = pygame.mixer.Sound("Sonidos/Imprimir.wav")
imprimir.set_volume(0.3)
#Otros
equis = pygame.image.load("Imagenes/Equis.png").convert()
equis.set_colorkey([0,0,0])
alerta = pygame.image.load("Imagenes/Alerta.png").convert()
alerta.set_colorkey([0,0,0])
#Hitbox botones menu
hit_start = pygame.Rect(372, 67, 145, 80)
hit_español = pygame.Rect(368,200,150,73)
hit_english = pygame.Rect(275,300,150,73)
hit_admin = pygame.Rect(207,435, 150, 72)
hit_ojo = pygame.Rect(350,551,50,35)
#hitbox botones maquina
hit_calidad = pygame.Rect(507,66, 59, 59)
hit_tipo = pygame.Rect(134,66,59,59)
hit_enter = pygame.Rect(301,253,97,50)
tipo = 1
calidad = 1
font_pantalla = pygame.font.SysFont("OCR A EXTENDED", 25)
#hitbox botones admin
hit_reset = pygame.Rect(207,565, 150, 73)
hit_shut_down = pygame.Rect(368,67,150,73)
hit_report = pygame.Rect(315,275,150,73)
#Volver
hit_volver = pygame.Rect(615,11, 70,70)
#Variables de entry
active = False
text = ""
font = pygame.font.Font(None, 35)
color_inactive = pygame.Color(45,154,212)
color_active = pygame.Color(47,117,161)
color = color_inactive
input_box = pygame.Rect(167, 550, 170, 40)
escenario = 1
codificado=""
privacidad = True
#Variables botones
idioma = "español"
press_a = False
press_s = False
incorrecto= False
press_rep = False
press_shut = False
press_res = False
press_vol = False
#Cargar imagenes de moneda
moneda1_10 = pygame.image.load("Imagenes/Moneda10/Moneda1.png").convert()
moneda1_10.set_colorkey([0,0,0])
moneda2_10 = pygame.image.load("Imagenes/Moneda10/Moneda2.png").convert()
moneda2_10.set_colorkey([0,0,0])
moneda3_10 = pygame.image.load("Imagenes/Moneda10/Moneda3.png").convert()
moneda3_10.set_colorkey([0,0,0])
moneda4_10 = pygame.image.load("Imagenes/Moneda10/Moneda4.png").convert()
moneda4_10.set_colorkey([0,0,0])
moneda5_10 = pygame.image.load("Imagenes/Moneda10/Moneda5.png").convert()
moneda5_10.set_colorkey([0,0,0])
moneda1_5 = pygame.image.load("Imagenes/Moneda5/Moneda1.png").convert()
moneda1_5.set_colorkey([0,0,0])
moneda2_5 = pygame.image.load("Imagenes/Moneda5/Moneda2.png").convert()
moneda2_5.set_colorkey([0,0,0])
moneda3_5 = pygame.image.load("Imagenes/Moneda5/Moneda3.png").convert()
moneda3_5.set_colorkey([0,0,0])
moneda4_5 = pygame.image.load("Imagenes/Moneda5/Moneda4.png").convert()
moneda4_5.set_colorkey([0,0,0])
moneda5_5 = pygame.image.load("Imagenes/Moneda5/Moneda5.png").convert()
moneda5_5.set_colorkey([0,0,0])
#Papel
papel = pygame.image.load("Imagenes/Hoja.png")
papel.set_colorkey([0,0,0])
#Listas de sprites
all_sprite_list = pygame.sprite.Group()
monedas_list = pygame.sprite.Group()
vuelto_list = pygame.sprite.Group()
#Variables monedas
cant_monedas = 0
agarrar = False
hit_entrada = pygame.Rect(66,547,68,11)
pago = 0
devolver = False
vuelto = 0
recogido =	True
"""Clase Moneda
#Atributos: image: imagen que enseña la instancia 
			rect: hitbox de la imagen
			frame: variable para elegir la imagen
			frames: lista de imagenes para la animacion
			valor: valor de la moneda 
			agarrar: boolean que verifica si la moneda esta agarrada o no
			animacion: boolean que verifica si la animacion esta en proceso

#Funciones
#change_frame(): obtiene el valor de self.frame
		#E:lista de imagenes #S:imagen segun el frame #R:-
#update():actualiza la imagen del sprite
		#E:instancia #S:muestra la imagen #R:- 
#getagarrar()
#setagarrar()
#getanimacion()
#setanimacion()
#getframe()
#getrect()
#setrect()
#getvalor()
"""
class Moneda(pygame.sprite.Sprite):
	def __init__(self, valor, frame):
		super().__init__()
		self.valor = valor
		self.frame = frame
		if valor == 5:
			self.frames = [moneda1_5, moneda2_5, moneda3_5, moneda4_5, moneda5_5]
			self.image = self.frames[self.frame]
		if valor == 10:
			self.frames = [moneda1_10, moneda2_10, moneda3_10, moneda4_10, moneda5_10]
			self.image = self.frames[self.frame]
		self.rect = moneda1_5.get_rect()
		self.agarrar = False
		self.animacion = False
	def getagarrar (self):
		return self.agarrar
	def setagarrar(self, value):
		self.agarrar = value
	
	def getanimacion (self):
		return self.animacion
	def setanimacion(self, value):
		self.animacion = value

	def getframe(self):
		return self.frame
	
	def getrect(self):
		return self.rect

	def setrect(self, x, y):
		self.rect.y = y
		self.rect.x = x

	def getvalor(self):
		return self.valor

	def change_frame(self, accion):
		if accion == "entrar":
			self.frame += 1
			if self.frame > (len(self.frames) - 1):
				self.frame = 0
			return self.frames[self.frame]
		
		if accion == "salir":
			self.frame -= 1
			return self.frames[self.frame]

	def update(self, accion):
		self.image = self.change_frame(accion)

"""Clase Dispensador
#Atributos: image: imagen que enseña la instancia 
			rect: hitbox de la imagen
			frame: variable para elegir la imagen
			sprites: lista de imagenes para la animacion 
			animacion: boolean que verifica si la animacion esta en proceso

#Funciones
#change_frame(): obtiene el valor de self.frame
		#E:lista de imagenes #S:imagen segun el frame #R:-
#update():actualiza la imagen del sprite
		#E:instancia #S:muestra la imagen #R:- 
#get_animacion()
#set_animacion()
#getframe()
"""
class Dispensador(pygame.sprite.Sprite):
	def __init__(self):
		super().__init__()
		self.image = dispensador1
		self.rect = (450,560)
		self.frame = 0
		self.sprites = [dispensador1,dispensador2,dispensador3,dispensador4,dispensador5,dispensador6,dispensador7,dispensador8]
		self.animacion = False 

	def get_animacion(self):
		return self.animacion

	def set_animacion(self):
		if self.animacion == True:
			self.animacion = False
		else:
			self.animacion = True

	def getframe(self):
		return self.frame	
			
	def change_frame(self):
		self.frame += 1
		if self.frame > (len(self.sprites)-1):
			self.frame = 0
		return self.sprites[self.frame]

	def update(self):
		self.image = self.change_frame()

#Creación del dispensador
dispenser_list = pygame.sprite.Group()
dispensador = Dispensador()
dispenser_list.add(dispensador)
all_sprite_list.add(dispensador)
#Variables de tabla
columna =0
fila = 0
orden = 2
#Variable de papel
mostrar = True
#main loop
while True:
	mouse= pygame.mouse.get_pos()
	for event in pygame.event.get():#Revisa los eventos
		if event.type == pygame.QUIT:  
			sys.exit()
		if escenario == 1: #Verifica el escenario
			if event.type == pygame.MOUSEBUTTONDOWN: 
				if input_box.collidepoint(mouse): #revisa si se da click en el entry
					active = True
					color = color_active
				else:
					active = False
					color = color_inactive
				if hit_english.collidepoint(mouse): #revisa si se da click en el boton de ingles
					beep.play()
					idioma = "english"
				if hit_español.collidepoint(mouse):#revisa si se da click en en el boton de españool
					beep.play()
					idioma= "español"
				if hit_start.collidepoint(mouse):#revisa si se da click en start
					beep.play()
					escenario = 2
					incorrecto = False
					codificado=""
				if hit_admin.collidepoint(mouse):#revisa si se da click en admin
					beep.play()
					if text == "1234": 
						escenario = 3
					else:
						text = ""
						codificado = ""
						incorrecto = True
				if hit_ojo.collidepoint(mouse): #revisa si se da click en el ojo
					if privacidad == True:
						privacidad = False
					elif privacidad == False:
						privacidad = True


			if event.type == pygame.MOUSEMOTION: #revisa si el mouse esta encima de los botonos para dar feedback
				if mouse[0]>=385 and mouse[0]<=506 and mouse[1]>=80 and mouse[1]<=140:
					press_s = True
				elif mouse[0]>=207 and mouse[0]<=357 and mouse[1]>=435 and mouse[1]<=507:
					press_a = True
				else:
					press_s =False
					press_a =False
			if event.type == pygame.KEYDOWN:
				if active: 
					if event.key == pygame.K_BACKSPACE: #detecta si es el boton es el de borrar para eliminar un caracter
						text = text [:len(text)-1]
						codificado = codificado[:len(text)-1]
					elif len(text)>6: #limite de caracteres
						if event.key == pygame.K_BACKSPACE: 
							text = text [:len(text)-1]
							codificado = modificado [:len(text)-1]
						else:
							None
					elif event.key == pygame.K_RETURN:
						beep.play()
						if text == "1234": 
							escenario = 3
						else:
							text = ""
							codificado=""
							incorrecto =True
					else: 
						codificado +="*"
						text += event.unicode
		
		if escenario == 2 and recogido == True: #verifica el escenario y estado
			if event.type == pygame.MOUSEBUTTONDOWN:
				if hit_tipo.collidepoint(mouse):#revisa si se da click el boton que cambia de tipo
					if agarrar == False:
						beep.play()
						if tipo == 3:
							tipo = 1
						elif tipo <3:
							tipo += 1
				
				if hit_calidad.collidepoint(mouse):#revisa si se da click el boton que cambia de calidad
					if agarrar == False:
						beep.play()
						if calidad == 3:
							calidad = 1
						elif calidad <3:
							calidad += 1

				if hit_enter.collidepoint(mouse): #revisa si se da click en el boton de enter
					if agarrar == False:
						beep.play()
					if pago>=precio and agarrar == False:
						dispensador.set_animacion()
						for m in monedas_list:
							monedas_list.remove(m)
							all_sprite_list.remove(m)
						if pago > precio:
							vuelto = pago - precio
							vuelto_tabla = vuelto
							devolver = True
						if pago == precio:
							vuelto_tabla = 0
						cant_monedas = 0
						pago_tabla = pago
						pago = 0

				if hit_volver.collidepoint(mouse): #revisa si se da click en el boton para volver al menu
					if agarrar == False:
						cerradura.play()
						for m in monedas_list:
							monedas_list.remove(m)
							all_sprite_list.remove(m)
						for v in vuelto_list:
							vuelto_list.remove(v)
							all_sprite_list.remove(v)
						privacidad = True
						cant_monedas = 0
						pago = 0
						agarrar = False
						escenario = 1
						devolver = False
						
				for m in monedas_list: 
					if m.getrect().collidepoint(mouse) and not m.getagarrar() and not agarrar and not devolver: #si da click sobre una moneda 
						m.setagarrar(True)
						agarrar =True
				for v in vuelto_list:
					if v.getrect().collidepoint(mouse): #si da click sobre el vuelto
						devolver = False
						monedas_list.remove(v)
						all_sprite_list.remove(v)

		if escenario == 2 and recogido == False: #verifica el escenario y estado
			if event.type == pygame.MOUSEBUTTONDOWN: #recoge el papel
				recogido = True
				mostrar = True
		
		if escenario == 3: #verifica el escenario
			if event.type == pygame.MOUSEBUTTONDOWN:	
				if hit_volver.collidepoint(mouse): #revisa si se da click en el boton para volver al menu
					beep.play()
					escenario = 1
					text = ""
					codificado = ""
					incorrecto = False
					privacidad = True
				if hit_report.collidepoint(mouse):#revisa si se da click en el boton para dar un reporte
					beep.play()
					matriz = [["Tipo","Codigo", "Mensaje", "Msg vendidos","Monto ventas"]]
					filas = []
					i =2
					crear = mensajes_s.cell(row = i, column = 6)
					#Creacion de matriz para la tabla de reportes
					while crear.value != None:
						if crear.value != 0:
							crear_tipo = mensajes_s.cell(row = i, column = 1)
							crear_codigo = mensajes_s.cell(row = i, column = 2)
							crear_msg = mensajes_s.cell(row = i, column = 3)
							valor_msg = crear_msg.value
							valor_msg = valor_msg[:20]
							crear_msg_vendidos = mensajes_s.cell(row = i, column = 6)
							crear_precio = mensajes_s.cell(row = i, column = 5)
							crear_monto = crear_msg_vendidos.value * crear_precio.value
							filas.append(crear_tipo.value)
							filas.append(crear_codigo.value)
							filas.append(valor_msg)
							filas.append(crear_msg_vendidos.value)
							filas.append(crear_monto)
							matriz.append(filas)
							filas = []
							i+=1
							crear = mensajes_s.cell(row = i, column = 6)
						else: 
							i+=1
							crear = mensajes_s.cell(row = i, column = 6)
					
					escenario = 4

				if hit_reset.collidepoint(mouse): #revisa si se da click en el boton reset
					beep.play()
					for i in range(2,52):
						borrar = mensajes_s.cell(row = i, column = 6, value= 0)
					i=2
					borrar_v = ventas_s.cell(row = i, column = 1)
					while borrar_v.value != None:
						for j in range(1,9):
							borrar_v = ventas_s.cell(row = i, column = j)
							borrar_v.value = None
						i+=1
						borrar_v = ventas_s.cell(row = i, column = 1)


					mensajes.save(filename = "Mensajes.xlsx")
					ventas.save(filename = "Ventas.xlsx")
				if hit_shut_down.collidepoint(mouse): #revisa si se da click en el boton shut down (apagar)
					beep.play()
					sys.exit()

			if event.type == pygame.MOUSEMOTION:#revisa si el mouse esta encima de los botonos para dar feedback
				if mouse[0]>=315 and mouse[0]<=465 and mouse[1]>=275 and mouse[1]<=348:
					press_rep = True
				elif mouse[0]>=207 and mouse[0]<=357 and mouse[1]>=565 and mouse[1]<=638:
					press_res = True
				elif mouse[0]>=368 and mouse[0]<=518 and mouse[1]>=67 and mouse[1]<=140:
					press_shut = True
				elif mouse[0]>=615 and mouse[0]<=685 and mouse[1]>=11 and mouse[1]<=81:
					press_vol = True
				else:
					press_rep =False
					press_res =False
					press_shut =False
					press_vol =False

		if escenario == 4:#verifica el escenario
			if event.type == pygame.MOUSEMOTION:#revisa si el mouse esta encima de los botonos para dar feedback
				if mouse[0]>=615 and mouse[0]<=685 and mouse[1]>=11 and mouse[1]<=81:
					press_vol = True
				else:
					press_vol =False
			if event.type == pygame.MOUSEBUTTONDOWN:	
				if hit_volver.collidepoint(mouse):#revisa si se da click en el boton para volver al modo administrador
					beep.play()
					escenario = 3
	#Se muestra en pantalla distintos elementos segun la variable escenario		
	if escenario == 1:
		tiempo =15
		for m in monedas_list:
			m.setagarrar(False)
		#Renderiza el fondo
		screen.blit(background,[0,0])
		#Input box
		pygame.draw.rect(screen, color, input_box, 4)
		#Renderiza el texto
		if privacidad == True:
			txt_surface = font.render(codificado, True, color_inactive)
		if privacidad == False:
			txt_surface = font.render(text, True, color_inactive)
		screen.blit(txt_surface, (input_box.x+5, input_box.y+7))
		#Botones
		if press_s == True:
			screen.blit(start_p,[372,67])
		if press_s ==False:
			screen.blit(start, [372,67])
		if press_a == True:
			screen.blit(admin_p,[207,435])
		if press_a == False:
			screen.blit(admin, [207,435])
		if incorrecto== True:
			screen.blit(equis, [250, 597])
		if idioma == "español":
			screen.blit(español_v, [368,200])
			screen.blit(english_r, [275,300])
		if idioma == "english":
			screen.blit(español_r, [368,200])
			screen.blit(english_v, [275,300])
		if privacidad == True:
			screen.blit(ojo_c, [350,545])
		if privacidad == False:
			screen.blit(ojo, [350,545])
	
	if escenario ==2:
		screen.blit(fondo,[0,0])
		tiempo = 10
		mouse_pos = pygame.mouse.get_pos()
		#cambio de texto en tipo dependiendo idioma
		if tipo == 1:
			if idioma == "español": 
				consejo = font_pantalla.render("Consejo", True,[255,255,255])
				screen.blit(consejo, (220,80))
			if idioma == "english":
				advice = font_pantalla.render("Advice", True, [255,255,255])
				screen.blit(advice, (220,80))
		if tipo == 2:
			if idioma == "español": 
				dicho = font_pantalla.render("Dicho", True,[255,255,255])
				screen.blit(dicho, (220,80))
			if idioma == "english":
				adage = font_pantalla.render("Adage", True, [255,255,255])
				screen.blit(adage, (220,80))
		if tipo == 3:
			if idioma == "español": 
				chiste = font_pantalla.render("Chiste", True,[255,255,255])
				screen.blit(chiste, (220,80))
			if idioma == "english":
				joke = font_pantalla.render("Joke", True, [255,255,255])
				screen.blit(joke, (220,80))
		#cambio de texto en calidad
		if calidad == 1:
			precio = 5 
			calidad_1 = font_pantalla.render("$"+str(precio), True, [255,255,255])
			screen.blit(calidad_1, (430,80))
		if calidad == 2:
			precio = 10
			calidad_2 = font_pantalla.render("$"+str(precio), True,[255,255,255])
			screen.blit(calidad_2, (430,80))
		if calidad == 3:
			precio = 15
			calidad_3 = font_pantalla.render("$"+str(precio), True,[255,255,255])
			screen.blit(calidad_3, (430,80))
		#Creaciond de monedas
		if cant_monedas == 0:
			moneda5 = Moneda(5,0)
			moneda5.setrect(80,620)
			monedas_list.add(moneda5)
			all_sprite_list.add(moneda5)
			moneda10 = Moneda(10,0)
			moneda10.setrect(200,620)
			monedas_list.add(moneda10)
			all_sprite_list.add(moneda10)
			cant_monedas+=2
		#Movimiento y animacion de las monedas
		for m in monedas_list:
			if m.getagarrar() == True:
				m.setrect(mouse_pos[0]-28, mouse_pos[1]-24)
				if hit_entrada.colliderect(m):
					m.setagarrar(False)
					m.setanimacion(True)
					agarrar =False

			if m.getanimacion() == True:
				if m.getframe() == 0:
					moneda_s.play()
				m.setrect(70,550)
				m.update("entrar")
				if m.getframe() == 0:
					m.setanimacion(False) 
					monedas_list.remove(m)
					all_sprite_list.remove(m)
					pago+=m.getvalor()
		#Animacion del dispensador
		for d in dispenser_list:
			if dispensador.get_animacion() == True:
				if m.getframe() == 0:
					imprimir.play()
				d.update()
				if d.getframe() == 0:
					d.set_animacion()
					recogido = False
					now = datetime.now()
		#Verifica si hay que dar vuelto y la cantidad a dar
		if devolver == True:
			if vuelto == 5:
				moneda5 = Moneda(5,4)
				moneda5.setrect(213,480)
				moneda5.setanimacion(True)
				vuelto_list.add(moneda5)
				all_sprite_list.add(moneda5)
				vuelto =0
			if vuelto == 10:
				moneda10 = Moneda(10,4)
				moneda10.setrect(213,480)
				moneda10.setanimacion(True)
				vuelto_list.add(moneda10)
				all_sprite_list.add(moneda10)
				vuelto =0
			else: 
				pass
		#Pago introducido por el usuario
		pago_intro = font_pantalla.render("$"+str(pago), True,[255,255,255])
		screen.blit(pago_intro, (330,170))
		#Dibuja los sprites
		all_sprite_list.draw(screen)
		#Animaciond el vuelto
		for v in vuelto_list:
			if v.getanimacion()== True:
				v.update("salir")
				if v.getframe()==0:
					v.setanimacion(False)
		#Seleccion de mensaje
		if recogido == 	False:
			if mostrar == True:
				#Eleccion segun idioma
				if idioma == "español":
					columna = 3
				else:
					columna = 4
				#Eleccion random de mensaje
				if tipo == 1:
					if calidad == 1:
						fila = random.randint(2,7)
					if calidad == 2:
						fila = random.randint(8,13)
					if calidad == 3:
						fila = random.randint(14,18)
				if tipo == 2:
					if calidad == 1:
						fila = random.randint(19,24)
					if calidad == 2:
						fila = random.randint(25,30)
					if calidad == 3:
						fila = random.randint(31,35)
				if tipo == 3:
					if calidad == 1:
						fila = random.randint(36,41)
					if calidad == 2:
						fila = random.randint(42,47)
					if calidad == 3:
						fila = random.randint(48,51)
				#Mensaje a enseñar y suma en numero de ventas
				mensaje = mensajes_s.cell(row = fila, column = columna)
				mensaje_valor = mensaje.value
				venta_m = mensajes_s.cell(row = fila, column = 6)
				venta_valor = venta_m.value
				venta_valor += 1
				venta_m = mensajes_s.cell(row = fila, column = 6,value = venta_valor)
				codigo = mensajes_s.cell(row = fila, column = 2)
				#Creacion de transaccion en el archivo de ventas
				n_trasnsaccion = ventas_s.cell(row = orden, column = 1)
				if n_trasnsaccion.value == None:
					n_trasnsaccion = ventas_s.cell(row = orden, column = 1, value= 1)
				else:
					while n_trasnsaccion.value != None:
						orden +=1
						n_trasnsaccion = ventas_s.cell(row = orden, column = 1)
					
					anterior= ventas_s.cell(row = orden-1, column = 1)
					anterior_valor = anterior.value
					n_trasnsaccion = ventas_s.cell(row = orden, column = 1, value= anterior_valor+1)
				if now.minute <10:
					minutos = "0"+str(now.minute)
				else: 
					minutos = str(now.minute)
				fecha = ventas_s.cell(row = orden, column = 2,value = str(now.day)+"/"+str(now.month)+"/"+str(now.year))
				hora = ventas_s.cell(row = orden, column = 3,value = str(now.hour)+":"+ minutos)
				tipo_v = ventas_s.cell(row = orden, column = 4,value = tipo)
				codigo_v = ventas_s.cell(row = orden, column = 5,value = codigo.value)
				monto = ventas_s.cell(row = orden, column = 6,value = precio)
				pago_v = ventas_s.cell(row = orden, column = 7,value = pago_tabla)
				vuelto_v = ventas_s.cell(row = orden, column = 8,value = vuelto_tabla)

				#Guardado de archivos
				mensajes.save(filename = "Mensajes.xlsx")
				ventas.save(filename = "Ventas.xlsx")
				mostrar = False
			#Mostrar el papel y el mensaje seleccionado
			screen.blit(papel,[100,100])
			mensajea = ""
			mensajeb = ""
			for i in range(len(mensaje_valor)):
				if mensaje_valor[i] == "/":
					mensajea = mensaje_valor[:i]
					mensajeb = mensaje_valor[i+1:]
					break
			if mensajeb == "":
				mensajea = mensaje_valor
				mensaje_text1 = excel_font.render(mensajea, True, [0,0,0])
				screen.blit(mensaje_text1,(130,300))
			else:
				mensaje_text1 = excel_font.render(mensajea, True, [0,0,0])	
				mensaje_text2 = excel_font.render(mensajeb, True, [0,0,0])
				screen.blit(mensaje_text1,(130,300))
				screen.blit(mensaje_text2,(130,350))

	if escenario ==3:
		#Fondo
		screen.blit(background,[0,0])
		#Triangulo de precaucion
		screen.blit(alerta,[207,420])
		#Botones
		if press_rep == False:
			screen.blit(report, [315,275])
		if press_rep == True:
			screen.blit(report_p, [315,275])
		if press_res == False:
			screen.blit(reset, [207,565])
		if press_res == True:
			screen.blit(reset_p, [207,565])
		if press_shut == False:
			screen.blit(shut_down, [368,67])
		if press_shut == True:
			screen.blit(shut_down_p, [368,67])	
		if press_vol == False:
			screen.blit(volver, [615,11])
		if press_vol == True:
			screen.blit(volver_p, [615,11])

	if escenario == 4:
		#Fondo
		screen.blit(fondo_admin,[0,0])
		#Botones
		if press_vol == False:
			screen.blit(volver, [615,11])
		if press_vol == True:
			screen.blit(volver_p, [615,11])
		x=20
		y=70
		#Renderizado del cuadro de reporte
		for i in range(len(matriz)):
			for j in range (len(matriz[0])):
				renderizar = excel_font.render(str(matriz[i][j]), True, [255,255,255])
				screen.blit(renderizar,(x,y))
				x += 100
				if j == 2 and i>0:
					x+=100
				if j == 2 and i==0:
					x+=50
				if j == 3 and i==0:
					x+=50

			y+= 50
			x= 20

	pygame.display.flip()
	clock.tick(tiempo)